from toolboxTMU import parameter, sqlLibrary, find_tap, initParameter, dataParser, convertBinList, signedInt16Handler
from openpyxl import Workbook
from pymodbus.client import ModbusSerialClient
import requests
import mysql.connector, time, datetime, math, openpyxl, sys, shutil, os
from requests.models import StreamConsumedError
from requests.exceptions import Timeout, RequestException

engineName = "Trafo X"
teleURL = 'http://192.168.4.120:1444/api/transformer/sendNotificationToTelegramGroup'
API_URL = "https://tmu.bambangdjaja.com/triggerAlarmNotification"
progStat = True
debugMsg = False
infoMsg = True
transmitterModeMinus = False

exhibitStat = False
OLTCstat = False
pressureStat = True
tempStat = True

companyKey = "P66geqk4bYQuetarke2Z"
raspiSerialNo = "1000000024b2178e"

def main():
    if infoMsg == True: print("1D|Initialize Program") 
    dataLen = 56
    watchedData = 29
    cycleTime = 2 / 60
    
    db = mysql.connector.connect(
        host = "localhost",
        user = "client",
        passwd = "raspi",
        database= "iot_trafo_client")
    cursor = db.cursor()

    client = ModbusSerialClient(method='rtu', port='/dev/ttyACM0', baudrate=9600)

    #init logger rawdata
    ts = time.strftime("%Y%m%d")
    pathStr = r'/home/pi/tmu/tmu-app-client-deploy/assets/datalog/rawdata/datalogger-'
    #pathStr = r'/home/pi/tmu-v2-smart/assets/rawdata-test/datalogger-'
    pathDatLog = pathStr + ts + '.xlsx'
    sheetName = ["Harmonic_phR", "Harmonic_phS", "Harmonic_phT"]
    pathBkup = r'/home/pi/tmu-v2-smart/assets/rawdata-test/backup/datalogger-backup-'
    pathDatBkup = pathBkup + ts + '.xlsx'
     
    try:
        wb = openpyxl.load_workbook(pathDatLog)
        if infoMsg == True: print("1D|Open Existing Excel")
    except:
        #create new datalog
        workbook = Workbook()
        workbook.save(pathDatLog)
        #create datalog's header
        wb = openpyxl.load_workbook(pathDatLog)
        sheet = wb.active
        sheet.title = "Raw_data"
        name = (('timestamp', 
                    'V-un', 'V-vn', 'V-wn', 'V-uv', 'V-vw', 'V-uw',
                    'I-u', 'I-v', 'I-w', 'Iavg', 'In',
                    'THDV-u', 'THDV-v', 'THDV-w', 'THDI-u', 'THDI-v', 'THDI-w',
                    'P-u', 'P-v', 'P-w', 'Ptot',
                    'Q-u', 'Q-v', 'Q-w', 'Q-tot',
                    'S-u', 'S-v', 'S-w', 'S-tot',
                    'PF-u', 'PF-v', 'PF-w', 'PFavg', 'Freq', 'kWh', 'kVARh',
                    'BusTemp-u', 'BusTemp-v', 'BusTemp-w', 'OilTemp',
                    'WTITemp-u', 'WTITemp-v', 'WTITemp-w',  'Press', 'Level',
                    'KRated-u', 'Derating-u', 'KRated-v', 'Derating-v', 'KRated-w', 'Derating-w',
                    'H2ppm', 'Moistppm', 'Vdiff-uv', 'Vdiff-vw', 'Vdiff-uw',
                    'trafoStatus', 'DIstat', 'DOstat', 'Alarm', 'Trip1', 'Trip2', 'Tap Position'),)
        for row in name:
            sheet.append(row)
        for member in sheetName:
            wb.create_sheet(member)
        for name in sheetName:
            sheetHarm = wb[name]
            rows = (('timestamp', 'V 1st', 'V 3rd' , 'V 5th' , 'V 7th' , 'V 9th' , 'V 11th' , 'V 13th' , 'V 15th' ,
                    'V 17th' , 'V 19th' , 'V 21st' , 'V 23rd' , 'V 25th' , 'V 27th' , 'V 29th' , 'V 31st',
                    'I 1st', 'I 3rd' , 'I 5th' , 'I 7th' , 'I 9th' , 'I 11th' , 'I 13th' , 'I 15th' ,
                    'I 17th' , 'I 19th' , 'I 21st' , 'I 23rd' , 'I 25th' , 'I 27th' , 'I 29th' , 'I 31st'),)
            for row in rows:
                sheetHarm.append(row)
        wb.save(pathDatLog)
        if debugMsg == True: print("1D|Create New Excel")
    
    inputData = [0]*dataLen
    currentStat = [0]*watchedData
    currentTrip = [0]*watchedData
    dataName = ['']*watchedData
    activeParam = [None]*watchedData
    activeFailure = [None]*watchedData
    dataSet = [parameter("Name", 0, False, None, None, None, None, 3, 0)]
    for i in range(0, dataLen-1):
        dataSet.append(parameter("Name", 0, False, None, None, None, None, 3, 0))
    messageReason = ['Extreme Low',
                'Low', 
                'Back Normal', 
                'High', 
                'Extreme High']
    msgEvent = [None] * watchedData
    msgAPI = [None] * watchedData
    msgReminder = [None] * watchedData
    telePrevTime = excelPrevTime  = excelSavePrevTime = datetime.datetime.now()
    cursor.execute(sqlLibrary.sqlFailure)
    listFailure = cursor.fetchall()
    for i in range(0, len(listFailure)):
        if listFailure[i][2] == None:
            activeFailure[activeFailure.index(None)] = listFailure[i]
    
    if infoMsg == True: print("1D|Start Loop")
    while progStat:
        if debugMsg == True: print("1D|1 Fetch DB Data")
        start_time = time.time()
        cursor.execute(sqlLibrary.sqlTrafoSetting)
        trafoSetting = cursor.fetchall()[0]
        cursor.execute(sqlLibrary.sqlTrafoData)
        trafoData = cursor.fetchall()[0]
        cursor.execute(sqlLibrary.sqlTripSetting)
        tripSetting = cursor.fetchall()[0]
        cursor.execute(sqlLibrary.sqlDIscan)
        inputIO = cursor.fetchall()
        cursor.execute(sqlLibrary.sqlDOscan)
        outputIO = cursor.fetchall()
        cursor.execute(sqlLibrary.sqlTrafoStatus)
        prevStat = list(cursor.fetchall()[0][1:])
        cursor.execute(sqlLibrary.sqlTripStatus)
        prevTrip = list(cursor.fetchall()[0][1:])
        db.commit()

        if debugMsg == True: print("1D|3 Import Active Failure")
        if len(activeFailure):
            for i in range(0, len(activeFailure)):
                if activeFailure[i]:
                    activeParam[i] = activeFailure[i][4]

        if debugMsg == True: print("1D|4a Read Modbus Slave")
        time.sleep(0.2)
        getDMCR = client.read_holding_registers(0x1301, 4, slave = 1)
        try:
            DMCRval = getDMCR.registers()
        except:
            DMCRval = [0, 0, 0, 4]
        
        if debugMsg == True: print("1D|5 Read Input IO")
        oilLevelAlarm = 1 if DMCRval[3]<=3 else 0
        oilLevelTrip = 1 if DMCRval[3]==1 else 0
        analogIn1 = (signedInt16Handler(DMCRval[1]))/10
        analogIn2 = (signedInt16Handler(DMCRval[0]))/1000

        if (oilLevelAlarm and oilLevelTrip) or oilLevelTrip:
            oilStat = 1
        elif oilLevelAlarm:
            oilStat = 2
        elif oilLevelAlarm == 0 and oilLevelTrip == 0:
            oilStat = 3
        inputData[44] = oilStat     #Oil Level

        if tempStat :
            inputData[39] = analogIn1
        else : 
            inputData[39] = 0

        if pressureStat:
            inputData[43] = analogIn2
        else:
            inputData[43] = 0

        if debugMsg == True: print("1D|6 Calculate WTI")
        for i in range(0, 3):
            inputData[i + 40] = inputData[39]
            
        if debugMsg == True: print("1D|9 Input all data DB")
        dataResult = initParameter(dataSet, inputData, trafoSetting, trafoData, tripSetting, dataLen) 
        sendData = [datetime.datetime.now()] + inputData
        cursor.execute(sqlLibrary.sqlInsertData, sendData)
        db.commit()
        if debugMsg == True: print("1D|10 Check Failures Stat")
        maxStat = 0
        i =  0
        for data in dataResult:
            if data.isWatched:
                maxStat = data.trafoStat if data.trafoStat > maxStat else maxStat
                currentStat[i] = data.status
                currentTrip[i] = data.trafoStat
                dataName[i] = data.name
                if data.status != prevStat[i]:
                    if data.status != 3:
                        if data.name in activeParam:
                            lastTimestamp = activeFailure[activeParam.index(data.name)][1]
                            duration = int((datetime.datetime.now() - lastTimestamp).total_seconds())
                            errorVal = [duration, activeFailure[activeParam.index(data.name)][0]]
                            cursor.execute(sqlLibrary.sqlResolveFailure, errorVal)
                            activeFailure[activeParam.index(data.name)] = None
                            activeParam[activeParam.index(data.name)] = None
                        errorVal = [datetime.datetime.now(), messageReason[data.status - 1], data.name, str(data.value)]
                        cursor.execute(sqlLibrary.sqlInsertFailure, errorVal)
                        cursor.execute(sqlLibrary.sqlLastFailure)
                        lastActive = cursor.fetchall()[0]
                        activeFailure[activeFailure.index(None)] = lastActive
                        loadProfile = str((round((data.value / trafoData[6]) * 10000))/100) + " Percent , Rated Current = " + str(trafoData[6])
                        msgEvent[i] = str(data.name + " " + messageReason[data.status - 1] + " , Value = " + (loadProfile if i == 3 or i == 4 or i == 5 else str(data.value)) + "\n" + "Time Occurence : " + str(datetime.datetime.now()))
                        eventType = "alarm" if data.status in [2, 4] else "trip"
                        eventValue = loadProfile if i == 3 or i == 4 or i == 5 else str(data.value)
                        msgAPI[i] = {
                            "companyKey": companyKey,
                            "raspiSerialNo": raspiSerialNo,
                            "time_start": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "failure_type": messageReason[data.status - 1],
                            "parameter": data.name,
                            "parameterValue": eventValue,
                            "duration": 0,
                            "event_type": eventType
                        }
                    elif data.status == 3:
                        lastTimestamp = activeFailure[activeParam.index(data.name)][1]
                        duration = int((datetime.datetime.now() - lastTimestamp).total_seconds())
                        errorVal = [duration, activeFailure[activeParam.index(data.name)][0]]
                        cursor.execute(sqlLibrary.sqlResolveFailure, errorVal)
                        activeFailure[activeParam.index(data.name)] = None
                        activeParam[activeParam.index(data.name)] = None
                        msgEvent[i] = None
                        msgAPI[i] = None
                i = i + 1
        if debugMsg == True: print("1D|11 Check state changes - send API & Tele")
        if prevStat != currentStat or prevTrip != currentTrip:
            for payload in msgAPI:
                if payload == None:
                    continue
                try:
                    response = requests.post(API_URL, json=payload, timeout=2)
                    response.raise_for_status()
                    if infoMsg == True: print("1D|%s" % response.json())
                    if infoMsg == True: print("1D|Alert sent to API successfully")
                except Timeout:
                    if infoMsg == True: print("1D|e: API Message Timeout")
                except RequestException as e:
                    if infoMsg == True: print("1D|%s" % Argument)
                    if infoMsg == True: print("1D|e: API Message Error")
            tele = list(filter(None, msgEvent))
            if tele:
                for message in tele:                
                    messages = engineName + " Says : " + "\n" + message
                    pload = {'message':messages}
                    try:
                        r = requests.post(teleURL, data = pload, timeout = 5, verify = False)
                    except Timeout:
                        if infoMsg == True: print("1D|e: Telegram Message Timeout")
                    except Exception as Argument:
                        if infoMsg == True: print("1D|%s" % Argument)
                        if infoMsg == True: print("1D|e: Telegram Message Error")
            else:
                pass
            cursor.execute(sqlLibrary.sqlUpdateTransformerStatus, currentStat)
            cursor.execute(sqlLibrary.sqlUpdateTripStatus, currentTrip)
            cursor.execute(sqlLibrary.sqlUpdateTrafoStat, (maxStat,))
            db.commit()
        else:
            pass
        binList = convertBinList(inputIO, outputIO, currentTrip)
        if int((datetime.datetime.now() - telePrevTime).total_seconds()) > 3600:
            if debugMsg == True: print("1D|12 Routine remind Tele")
            #print("sekadar mengingatkan")
            for i in range(0, len(activeFailure)):
                if activeFailure[i]:
                    failureIndex = dataName.index(activeFailure[i][4])
                    msgReminder[failureIndex] = str(activeFailure[i][4] + " " + activeFailure[i][3] + " , Value = " + activeFailure[i][5] + "\n" + "Time Occurence : " + str(activeFailure[i][1]))                    
                    messages = engineName + " Says : " + "\n" + msgReminder
                    pload = {'message':messages}
                    try:
                        r = requests.post(teleURL, data = pload, timeout = 5, verify = False)
                    except Timeout:
                        if infoMsg == True: print("1D|e: Telegram Message Timeout")
                    except Exception as Argument:
                        if infoMsg == True: print("1D|e: Telegram Message Error")

            telePrevTime = datetime.datetime.now()
        #print(inputData)
        if int((datetime.datetime.now() - excelPrevTime).total_seconds()) > 3:
            if debugMsg == True: print("1D|12A Routine Add data to work stage excel")
            sendLog = [datetime.datetime.now().strftime("%H:%M:%S")] + inputData + [maxStat] + binList + [OLTCstat]
            sendLog = ((tuple(sendLog)),)
            sheet = wb["Raw_data"]
            for row in sendLog:
                sheet.append(row)
            excelPrevTime = datetime.datetime.now()
        if int((datetime.datetime.now() - excelSavePrevTime).total_seconds()) > 180:
            if debugMsg == True: print("1D|12B Routine Save Excel")
            if infoMsg == True: print("1D|Check Current Excel Size")
            if os.path.isfile(pathDatBkup) and os.path.getsize(pathDatBkup) >= os.path.getsize(pathDatLog):
                if infoMsg == True: print("1D|Excel Smaller than backup, replacing")
                shutil.copy2(pathDatBkup, pathDatLog)
            else:
                #create backup
                if infoMsg == True: print("1D|Backup Excel")
                shutil.copy2(pathDatLog, pathDatBkup)
            #print("save excel data here")
            try:
                if infoMsg == True: print("1D|Try to save Excel from work stage")
                wb.save(pathDatLog)
                time.sleep(0.5)
                if infoMsg == True: print("1D|Excel Size %s " % (os.path.getsize(pathDatLog)))
                if infoMsg == True: print("1D|Backup Size %s " % (os.path.getsize(pathDatBkup)))
                if (os.path.getsize(pathDatBkup) - os.path.getsize(pathDatLog)) < 3000:
                    if infoMsg == True: print("1D|Save Success")
                else:
                    raise Exception("backup larger than saved excel")
            except Exception as e:
                if infoMsg == True: print("1D|%s" % e)
                if infoMsg == True: print("1D|e: Save Failed, return to backup, restart system")
                shutil.copy2(pathDatBkup, pathDatLog)
                if infoMsg == True: print("1D|Restart")
            excelSavePrevTime = datetime.datetime.now()
                        
        cycleTime = (round(10000 * (time.time() - start_time)))/10000
        if debugMsg == True: print("1D|Cycle time %s" % cycleTime)
        print("1T|%s" % datetime.datetime.now())
        sys.stdout.flush()
        time.sleep(4)
        
if __name__ == "__main__":
    main()
